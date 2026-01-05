# concrete_analysis_expanded.py

"""
Expanded analysis of UCI Concrete Compressive Strength dataset.
Includes:
- Calculated water/cement ratio
- Correlation analysis
- Styled Excel export
"""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import requests
import os
import sys

# ----------------------------
# 1. Load and preprocess
# ----------------------------
url = "https://archive.ics.uci.edu/ml/machine-learning-databases/concrete/compressive/Concrete_Data.xls"

# URL of the .xls file
url = "https://archive.ics.uci.edu/ml/machine-learning-databases/concrete/compressive/Concrete_Data.xls"
local_filename = "Concrete_Data.xls"

# Download using requests
response = requests.get(url)
with open(local_filename, "wb") as f:
    f.write(response.content)

os.system("in2csv Concrete_Data.xls > Concrete_Data.csv")

df = pd.read_csv("Concrete_Data.csv")

df.columns = [c.strip().lower().split("(")[0].replace(" ", "_").rstrip("_") for c in df.columns]
print("Dataset loaded with shape:", df.shape)
print("Columns:", df.columns)

# Calculate water/cement ratio
df["w_c_ratio"] = df["water"] / df["cement"]

# ----------------------------
# 2. Grouping and stats
# ----------------------------
summary = df.groupby("age")[["concrete_compressive_strength", "w_c_ratio"]].agg(["mean", "std"])

# ----------------------------
# 3. Plot: Strength vs. Age and w/c ratio
# ----------------------------
plt.figure(figsize=(10, 5))
plt.scatter(df["w_c_ratio"], df["concrete_compressive_strength"], alpha=0.5)
plt.xlabel("Water/Cement Ratio")
plt.ylabel("Compressive Strength (MPa)")
plt.title("Concrete Strength vs. Water/Cement Ratio")
plt.grid(True)
plt.tight_layout()
plt.savefig("strength_vs_wc.png")
plt.show()

# ----------------------------
# 4. Styled Excel export
# ----------------------------
wb = Workbook()
ws1 = wb.active
ws1.title = "Summary by Age"

# Write summary with styles
for r_idx, row in enumerate(dataframe_to_rows(summary, index=True, header=True), 1):
    ws1.append(row)
    if r_idx == 1:
        for cell in ws1[r_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Add raw data to second sheet
ws2 = wb.create_sheet(title="Raw Data")
for row in dataframe_to_rows(df, index=False, header=True):
    ws2.append(row)

# Save
wb.save("concrete_summary_styled.xlsx")