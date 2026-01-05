import sys
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np

from weather_data import *

# Ladda ner värden från olika väderstationer

station_ids = [
    'USW00094728_NYC',  # New York City
    'USW00012839_MIA',  # Miami
    'USW00023062_DEN'   # Denver
]

# Ange start- och slutdamm för datan

start_date = '2020-01-01'
end_date = '2023-12-31'

# Ladda ner och rensa väderdata

print("Laddar ner väderdata...\n")
weather_df = download_weather_data(station_ids, start_date, end_date)
weather_df.to_csv('weather_data.csv', index=False)

print()
print(f"Laddat ner {len(weather_df)} rader.\n")
print("Första raderna av väderdata:\n")
print(weather_df.head(10))
print()

print("Data typer i weather_df:")
print(weather_df.dtypes)
print()
print(f"Form på weather_df: {weather_df.shape}")

# ------------------------------------------
# Pivot tabell  - Temperatur per stad och år
# ------------------------------------------

print("\nGenomsnittlig temperatur per stad och år")

temp_pivot = weather_df.pivot_table(
    values='TAVG', 
    index='YEAR', 
    columns='CITY', 
    aggfunc='mean'
).round(1)

print(temp_pivot)

# -------------------------------------------------------------
# Gruppera efter månad och stad för att beräkna klimatstatistik
# -------------------------------------------------------------

print("\n")
print("Gruppera efter månad och stad för att beräkna klimatstatistik\n")

monthly_stats_mean = weather_df.groupby(['CITY']).mean(numeric_only=True).round(2)
print(monthly_stats_mean)
print()

monthly_stats_aggregate = weather_df.groupby(['CITY', 'MONTH']).agg({
    'TAVG': ['mean', 'std'],
    'PRCP': ['mean', 'sum'],
    'TMAX': 'max',
    'TMIN': 'min',
    'AWND': 'mean'
}).round(2)

# Platta kolumnnamn för bättre läsbarhet

#monthly_stats.columns = ['_'.join(col).strip() for col in monthly_stats.columns]

#print(monthly_stats.head(10))
print(monthly_stats_aggregate)
print()

yearly_stats_aggregate = weather_df.groupby(['CITY', 'YEAR']).agg({
    'TAVG': ['mean', 'std'],
    'PRCP': ['mean', 'sum'],
    'TMAX': 'max',
    'TMIN': 'min',
    'AWND': 'mean'
}).round(2)

print(yearly_stats_aggregate)

# -------------------------------------------
# Slå samman - Lägg till klimatzoninformation
# -------------------------------------------

print("\n")
print("Slå samman data med klimatzoninformation\n")

# Skapa en DataFrame med klimatzoninformation
# Denna data skulle normalt komma från en extern källa eller API

climate_zones = pd.DataFrame({
    'CITY': ['NYC', 'MIA', 'DEN'],
    'CLIMATE_ZONE': ['Humid Continental', 'Tropical', 'Semi-Arid'],
    'ELEVATION_FT': [33, 6, 5280],
    'LATITUDE': [40.7, 25.8, 39.7],
    'LONGITUDE': [-74.0, -80.3, -105.0]
})

print("Klimatzoninformation:\n")
print(climate_zones)
print()

# Slå samman väderdata med klimatzoninformation

weather_enhanced = weather_df.merge(climate_zones, on='CITY', how='left')
print("\nFörsta raderna av väderdata med klimatzoninformation:\n")
print(weather_enhanced)

#print(weather_enhanced[['CITY', 'CLIMATE_ZONE', 'ELEVATION_FT', 'TAVG']].drop_duplicates())

# ----------------------------------------------------
# Kombinera årliga summeringar för varje stad
# Detta skulle normalt vara en del av en större analys
# ----------------------------------------------------

print("\n")
print("Årliga summeringar per stad\n")

annual_summaries = []
for year in weather_df['YEAR'].unique():
    year_data = weather_df[weather_df['YEAR'] == year]
    
    summary = year_data.groupby('CITY').agg({
        'TAVG': 'mean',
        'PRCP': 'sum',
        'TMAX': 'max',
        'TMIN': 'min',
        'AWND': 'mean'
    }).round(2)
    
    summary['YEAR'] = year
    summary['SEASON_TEMP_RANGE'] = summary['TMAX'] - summary['TMIN']

    print(summary)
    annual_summaries.append(summary)

# Concatenate all annual summaries
combined_annual = pd.concat(annual_summaries, ignore_index=False)
print(combined_annual)

# -------------------------------------------------------
# Visualisering av Pandas data med Matplotlib och Seaborn
# -------------------------------------------------------

# Importera matplotlib och seaborn

import matplotlib.pyplot as plt
import seaborn as sns

# Definiera stil på uppritning

plt.style.use('seaborn-v0_8')
sns.set_palette("husl")

# -------------------------------------
# Visualisering tidsserier (Matplotlib)
# -------------------------------------

plt.figure()
for city in weather_df['CITY'].unique():
    city_data = weather_df[weather_df['CITY'] == city]
    plt.plot(city_data['DATE'], city_data['TAVG'], label=city, linewidth=2, alpha=0.8)

plt.title('Temperaturtrender över tid', fontsize=14, fontweight='bold')
plt.xlabel('Damm')
plt.ylabel('Medeltemperatur (°C)')
plt.legend()
plt.xticks(rotation=45)
plt.grid(True, alpha=0.3)


# -----------------------------------------------
# Låddiagram av temperatur distribution (boxplot)
# -----------------------------------------------

plt.figure()
sns.boxplot(data=weather_df, x='CITY', y='TAVG')
plt.title('Temperaturfördelning per stad', fontsize=14, fontweight='bold')
plt.ylabel('Medeltemperatur (°C)')

# -------------------------------------------
# Heatmap över temperaturmönster över månaden
# -------------------------------------------

plt.figure()

monthly_temp_pivot = weather_df.pivot_table(
    values='TAVG', index='MONTH', columns='CITY', aggfunc='mean'
)

sns.heatmap(monthly_temp_pivot, annot=True, cmap='RdYlBu_r', fmt='.1f', cbar_kws={'label': 'Temperatur (°C)'})
plt.title('Månatliga temperaturmönster', fontsize=14, fontweight='bold')
plt.ylabel('Månad')

# ------------------------------------------
# Scatter plott för temperatur och nederbörd
# ------------------------------------------

plt.figure()

for city in weather_df['CITY'].unique():
    city_data = weather_df[weather_df['CITY'] == city]
    plt.scatter(city_data['PRCP'], city_data['TAVG'], label=city, alpha=0.7, s=50)

plt.title('Temperatur vs Nederbörd', fontsize=14, fontweight='bold')
plt.xlabel('Nederbörd (mm)')
plt.ylabel('Medeltemperatur (°C)')
plt.legend()
plt.grid(True, alpha=0.3)

# ----------------------------------------
# Stapeldiagram över total årlig nederbörd
# ----------------------------------------

plt.figure()

annual_precip = weather_df.groupby(['YEAR', 'CITY'])['PRCP'].sum().unstack()
annual_precip.plot(kind='bar', ax=plt.gca())

plt.title('Total årlig nederbörd', fontsize=14, fontweight='bold')
plt.ylabel('Nederbörd (mm)')
plt.xticks(rotation=45)


# ----------------------------------------
# Violinplot över vindhastighetsfördelning
# ----------------------------------------

plt.figure()

sns.violinplot(data=weather_df, x='CITY', y='AWND')
plt.title('Vindhastighetsfördelning', fontsize=14, fontweight='bold')
plt.ylabel('Vindhastighet (mph)')


# ----------------------------
# Linjeplot med felmarkeringar
# ----------------------------

plt.figure()

monthly_avg = weather_df.groupby(['MONTH', 'CITY'])['TAVG'].agg(['mean', 'std']).reset_index()

for city in weather_df['CITY'].unique():
    city_monthly = monthly_avg[monthly_avg['CITY'] == city]
    plt.errorbar(city_monthly['MONTH'], city_monthly['mean'], 
                yerr=city_monthly['std'], label=city, capsize=5, linewidth=2, alpha=0.8)

plt.title('Månadstemperatur med variation', fontsize=14, fontweight='bold')
plt.xlabel('Månad')
plt.ylabel('Temperatur (°C)')
plt.legend()
plt.grid(True, alpha=0.3)

# ----------------------------------------
# Korrelations-heatmap över vädervariabler
# ----------------------------------------

plt.figure()

correlation_data = weather_df[['TAVG', 'TMAX', 'TMIN', 'PRCP', 'AWND']].corr()

sns.heatmap(correlation_data, annot=True, cmap='coolwarm', center=0, fmt='.2f',
           cbar_kws={'label': 'Korrelationskoefficient'})

plt.title('Korrelationsmatris för vädervariabler', fontsize=14, fontweight='bold')

# ----------------------------------------------------------
# Staplat stapeldiagram över säsongsvariationer i temperatur
# ----------------------------------------------------------

plt.figure()

weather_df['SEASON'] = weather_df['MONTH'].map({
    12: 'Vinter', 1: 'Vinter', 2: 'Vinter',
    3: 'Vår', 4: 'Vår', 5: 'Vår',
    6: 'Sommar', 7: 'Sommar', 8: 'Sommar',
    9: 'Höst', 10: 'Höst', 11: 'Höst'
})

seasonal_range = weather_df.groupby(['CITY', 'SEASON']).agg({
    'TMAX': 'mean',
    'TMIN': 'mean'
}).reset_index()

seasonal_range['TEMP_RANGE'] = seasonal_range['TMAX'] - seasonal_range['TMIN']
seasonal_pivot = seasonal_range.pivot(index='CITY', columns='SEASON', values='TEMP_RANGE')
seasonal_pivot[['Vinter', 'Vår', 'Sommar', 'Höst']].plot(kind='bar', stacked=True, ax=plt.gca())

plt.title('Säsongsvariationer i temperaturspann', fontsize=14, fontweight='bold')
plt.ylabel('Temperaturspann (°C)')
plt.xticks(rotation=45)

plt.show()
sys.exit(0)

print("\nAll plots have been generated successfully!")

# =============================================================================
# 4. EXPORTING DATA TO EXCEL (SIMPLE AND FORMATTED)
# =============================================================================

print("\n" + "="*60)
print("EXCEL EXPORT EXAMPLES")
print("="*60)

# 4.1 Simple Excel Export
print("\n4.1 Simple Excel Export")

# Create a simple export with multiple sheets
with pd.ExcelWriter('weather_data_simple.xlsx', engine='openpyxl') as writer:
    # Raw data sheet
    weather_df.to_excel(writer, sheet_name='Raw_Data', index=False)
    
    # Monthly statistics sheet
    monthly_stats.to_excel(writer, sheet_name='Monthly_Stats')
    
    # Annual summaries sheet
    combined_annual.to_excel(writer, sheet_name='Annual_Summary')
    
    # Climate zones reference
    climate_zones.to_excel(writer, sheet_name='Climate_Zones', index=False)

print("Simple Excel file 'weather_data_simple.xlsx' created successfully!")

# 4.2 Formatted Excel Export
print("\n4.2 Formatted Excel Export with Styling")

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, Reference
import openpyxl

# Create formatted Excel file
with pd.ExcelWriter('weather_data_formatted.xlsx', engine='openpyxl') as writer:
    
    # Export data to different sheets
    weather_df.to_excel(writer, sheet_name='Dashboard', index=False, startrow=4)
    temp_pivot.to_excel(writer, sheet_name='Temperature_Analysis')
    monthly_stats.to_excel(writer, sheet_name='Detailed_Stats')
    
    # Access the workbook and worksheets for formatting
    workbook = writer.book
    
    # Format Dashboard sheet
    dashboard = workbook['Dashboard']
    
    # Add title
    dashboard['A1'] = 'Weather Data Analysis Report'
    dashboard['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    dashboard['A1'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    dashboard.merge_cells('A1:G1')
    dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add subtitle
    dashboard['A2'] = f'Data Period: {start_date} to {end_date}'
    dashboard['A2'].font = Font(size=12, italic=True)
    dashboard.merge_cells('A2:G2')
    dashboard['A2'].alignment = Alignment(horizontal='center')
    
    # Add summary statistics
    dashboard['A3'] = f'Total Records: {len(weather_df)} | Cities: {", ".join(weather_df["CITY"].unique())}'
    dashboard['A3'].font = Font(size=10, color='666666')
    dashboard.merge_cells('A3:G3')
    dashboard['A3'].alignment = Alignment(horizontal='center')
    
    # Format headers (row 5, since data starts at row 5 due to startrow=4)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, column_title in enumerate(weather_df.columns, 1):
        cell = dashboard.cell(row=5, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for col_num in range(1, len(weather_df.columns) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        
        # Check header length
        header_cell = dashboard.cell(row=5, column=col_num)
        if header_cell.value:
            max_length = len(str(header_cell.value))
        
        # Check data length (sample first 100 rows for performance)
        for row_num in range(6, min(106, len(weather_df) + 6)):
            cell = dashboard.cell(row=row_num, column=col_num)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 20)
        dashboard.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders to data table
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    max_row = len(weather_df) + 5
    max_col = len(weather_df.columns)
    
    for row in dashboard.iter_rows(min_row=5, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row > 5:  # Data rows (not header)
                cell.alignment = Alignment(horizontal='center')
    
    # Format Temperature Analysis sheet
    temp_sheet = workbook['Temperature_Analysis']
    temp_sheet['A1'] = 'Temperature Analysis by City and Year'
    temp_sheet['A1'].font = Font(size=14, bold=True)
    temp_sheet['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    temp_sheet['A1'].font = Font(color='FFFFFF', bold=True)
    
    # Apply color scale to temperature data
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='87CEEB',  # Light blue for cold
        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow for moderate
        end_type='max', end_color='FF6B6B'  # Red for hot
    )
    
    # Apply color scale to the temperature pivot data
    data_range = f'B3:E{3 + len(temp_pivot)}'
    temp_sheet.conditional_formatting.add(data_range, color_scale_rule)

print("Formatted Excel file 'weather_data_formatted.xlsx' created successfully!")

# 4.3 Excel Export with Charts
print("\n4.3 Advanced Excel Export with Charts")

# This would require additional openpyxl chart creation
# For brevity, showing the concept with comments

"""
Advanced Excel export could include:
- Embedded charts (line charts, bar charts)
- Conditional formatting
- Data validation
- Formulas and calculations
- Multiple formatted tables
- Print settings and page layout
"""

# Create summary statistics for final display
print("\n" + "="*60)
print("ANALYSIS SUMMARY")
print("="*60)

print("\nClimate Summary by City:")
climate_summary = weather_df.groupby('CITY').agg({
    'TAVG': ['mean', 'min', 'max'],
    'PRCP': 'sum',
    'AWND': 'mean'
}).round(2)

climate_summary.columns = ['Avg_Temp', 'Min_Temp', 'Max_Temp', 'Total_Precip', 'Avg_Wind']
print(climate_summary)

print("\nExported Files:")
print("1. weather_data_simple.xlsx - Basic multi-sheet export")
print("2. weather_data_formatted.xlsx - Professionally formatted with styling")

print("\nKey Pandas Techniques Demonstrated:")
print("• Data download and ingestion from APIs")
print("• Pivot tables and cross-tabulation")
print("• GroupBy operations and aggregations")
print("• Data merging and joining")
print("• Time series analysis")
print("• Data visualization integration")
print("• Professional Excel export with formatting")

print(f"\nTotal execution completed successfully!")
print(f"Processed {len(weather_df)} weather records across {len(weather_df['CITY'].unique())} cities")